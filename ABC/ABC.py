import pandas as pd
from matplotlib.font_manager import FontProperties
import pandas as pd
from copy import deepcopy
from src_dir import res_initial
font_set = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=12)


import copy
from src_dir import fit_fun
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import numpy as np

def random_pick(probabilities):
    x = random.uniform(0, 1)
    cumulative_probability = 0.0
    for item_probability in probabilities:
        cumulative_probability += item_probability
        if x < cumulative_probability:
            return probabilities.index(item_probability)

def show_detail(res, id_teacher):
    '''
    :param res: 初始化的多个粒子
    :return: 以文字展示
    '''
    valid_count = 0
    valid_lz = []
    for r in res:
        flag = 0
        for j, k in enumerate(r):
            flg = 0
            for te in k[0].keys():
                for su in k[1]:
                    if (te == id_teacher[su]):
                        flg = 1
                        flag = 1

        if flag == 0:
            valid_count += 1
            valid_lz.append(r)


def plot_scale(lz,i,id_score):
    '''
    :param lz: 某个粒子
    :return:
    '''
    # fig, axes = plt.subplots(2, 3, figsize=(20, 10))
    # fig, axes = plt.subplots(2, 2, figsize=(20, 15))

    for index, id_group in enumerate(lz):
        if isinstance(id_group,dict):
            continue
        plt.figure()
        fig = plt.subplot()
        group_score = []
        for id_ in id_group[1]:
            group_score.append(id_score[id_])
        group_score = pd.Series(group_score)

        # axes[int(index / 2)][index % 2].hist(group_score, bins=[1, 1.5, 2, 2.5, 3, 3.5, 4],density=True)
        fig.hist(group_score, bins=[1, 1.5, 2, 2.5, 3, 3.5, 4],density=True)


        # axes[int(index / 2)][index % 2].hist(group_score, bins=[0, 2, 2.5, 2, 2.5, 4],density=True)
        # title = "第" + str(index) + "组"
        # axes[int(index / 2)][index % 2].set_title(u"第"+str(index+1)+u"组",fontproperties=font_set)

        fig.set_xlabel(u"绩点区间",fontproperties=font_set)

        # axes[int(index / 2)][index % 2].set_xlabel(u"绩点区间",fontproperties=font_set)

        fig.set_ylabel(u"频率/间距",fontproperties=font_set)
        # axes[int(index / 2)][index % 2].set_ylabel(u"频率/间距",fontproperties=font_set)

        plt.savefig("static/image/" + str(index) + "_07.jpg")
        plt.savefig("static/image/" + str(index) + "_07.svg",format="svg")


def plt_show_all(res, fit_list, n,id_score):
    '''
    :param res: res 为初始化的多个粒子
    :param fit: 为res中粒子的适应度
    :param n: 展示n个粒子
    :return:
    '''

    for i, r in enumerate(list(np.argsort(np.array(fit_list[:n])))):
        plot_scale(res[r],i,id_score)


def Diversity(lzs=[], n=0):
    diversity = 0
    i = 0
    while i < len(lzs):
        j = 0
        diversity_i = 0
        while j < len(lzs):
            if j == i:
                j += 1
                continue
            # print(i, j)
            sum = 0
            for group in range(n):
                # print(set(lzs[i][group][0]['teachers']).intersection(set(lzs[j][group][0]['teachers'])))
                len_max = max(len(lzs[i][group][0]['teachers']),len(lzs[i][group][0]['teachers']))
                sum += (len_max - len(set(lzs[i][group][0]['teachers']).intersection(set(lzs[j][group][0]['teachers'])))) / len_max
            sum = sum / n
            # print('两个粒子差异性', sum)
            diversity_i += sum
            j += 1
            if i == len(lzs):
                break
        diversity += diversity_i / (len(lzs)-1)
        # print('一个粒子与总体的差异性', diversity_i / (len(lzs)-1))
        i += 1
    # print("diversity", diversity/len(lzs))
    return diversity/len(lzs)


def begin(n=4, x=2, n_groups=50, teachers=4, accuracy_level=2, clash_teacher=[], same_teacher=[], rd=0, save=0.8):
    '''
        data '评阅数据集2.xlsx'为数据
        n为组数
        x为每组人数不同的程度
        n_groups为粒子数
        teachers 为答辩老师
    '''
    data = pd.read_excel(io='src_dir/new_data1.xlsx')
    data.columns = ["id", "score", "teacher"]
    # 学号（学生）对应的老师
    id_teacher = dict([*zip(data["id"], data["teacher"])])
    # （学号）学生对应的绩点
    id_score = dict([*zip(data["id"], data["score"])])
    score_scale = [0, 0, 0, 0, 0]
    for i in data["score"]:
        if i < 2.0:
            score_scale[0] += 1
        elif i < 2.5:
            score_scale[1] += 1
        elif i < 3.0:
            score_scale[2] += 1
        elif i < 3.5:
            score_scale[3] += 1
        else:
            score_scale[4] += 1
    # 以上是对id_teacher，id_score，score_scale的初始化
    for i in range(5):
        score_scale[i] = score_scale[i] / data.shape[0]
    # 初始化函数 参数具体含义在fit

    score_series = pd.Series(data["score"])
    # print(score_series)
    # axe = plt.subplot()
    # axe.hist(score_series, bins=[1, 1.5, 2, 2.5, 3, 3.5, 4],density=True)
    # axe.set_xlabel(u"绩点区间",fontproperties=font_set)
    # axe.set_ylabel(u"频率/间距",fontproperties=font_set)
    # axe.set_title(u"全局绩点分布",fontproperties=font_set)
    # axe.hist(score_series, bins=[0, 2, 2.5, 3, 3.5, 4],density=True)
    plt.savefig("static/image/global.jpg")
    plt.savefig("static/image/global.svg",format="svg")

    temp_lzcsh = res_initial.lzcsh(data, n, x, n_groups, teachers, clash_teacher, same_teacher)
    # print(temp_lzcsh)
    if isinstance(temp_lzcsh,int):
        # print("无法初始化")
        return -2
    else:
        lzs, teacher_statu = temp_lzcsh
    # res_his 记录历史最优
    # 初始化完的粒子是第一轮粒子的历史最优
    res_his = copy.deepcopy(lzs)
    cost = 0
    best_indx = 0
    fit = []
    fit_sum = 0
    probability = []
    change_count = []
    Diversity_list = []
    for lz in range(len(lzs)):
        change_count.append(0)
        temp_cost = fit_fun.fit_all(lzs[lz], id_teacher, id_score, score_scale, n, same_teacher)
        if temp_cost < 0:   # 删除适应度小于0的蜜源
            while True:
                temp_lz, t_s = res_initial.lzcsh(data, n, x, 1, teachers, clash_teacher, same_teacher)
                temp_cost = fit_fun.fit_all(temp_lz[0], id_teacher, id_score, score_scale, n, same_teacher)
                if temp_cost > 0:
                    lzs[lz] = temp_lz[0]
                    break
        fit.append(temp_cost)
        fit_sum += temp_cost

        if temp_cost > cost:    # 选最好的粒子
            cost = temp_cost
            best_indx = lz

    for lz in fit:
        # print(lz)
        probability.append(lz/fit_sum)
    # print(probability)
    pre_best = deepcopy(fit[best_indx])
    # print('pre best:',fit[best_indx])
    # print(best_indx)
    # for e in lzs[best_indx]:
    #     print(e)
    # g_best 全局最优粒子
    g_best = copy.deepcopy(lzs[best_indx])
    g_best_list = []
    advance = 0
    g_advance = 0
    # -----------------
    for iterm in range(200):  # 迭代150轮
        for lz in range(len(lzs)):
            lz_re = deepcopy(lzs[lz])
            # print("select",lz,fit[lz])
            teacher_set = set(id_teacher.values())
            for n_group in range(len(lzs[lz]) - 1):

                lz_tea_set = set(list(lzs[lz][n_group][0])[:-1])
                # 从当前粒子这组选一部分老师集合
                lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * save)))

                # 迭代后这组老师应该从这个集合中抽
                wait_selec_set_to_list = list(lz_tea_set)

                # 要与剩下老师集合做交集
                wait_selec_set_to_list = list(set(wait_selec_set_to_list) & teacher_set)

                random_tea = random.sample(list(teacher_set.difference(wait_selec_set_to_list)), round(len(lz_tea_set) * 0.5))
                wait_selec_set_to_list += random_tea

                # --dp  从n个数中抽 出x个数 使其和最接近m  关于第5点的启发式
                # 这里的n 是wait_selec_set_to_list 老师对应的学生数 m是平均每组学生数
                state = []
                for i in range(100):
                    state.append([])
                for i in state:
                    for j in range(50):
                        i.append(0)
                dp = []
                for i in range(100):
                    dp.append(0)

                value = []
                for tea in wait_selec_set_to_list:
                    value.append(len(teacher_statu[tea]))
                n2 = len(value)

                ave_stu = len(id_teacher) / len(g_best)

                sum_temp = 0
                m = round(ave_stu - sum_temp)

                for i in range(n2):
                    j = m
                    while j >= value[i]:
                        tmp = dp[j - value[i]] + value[i]
                        if tmp > dp[j]:
                            dp[j] = tmp
                            try:                        ###
                                state[i][j] = 1
                            except:
                                #print("出现异常，分组过少")
                                return -1
                        j -= 1

                i = n2
                j = m
                ans = []

                while i > 0:
                    i -= 1
                    if state[i][j] == 1:
                        ans.append(value[i])
                        j -= value[i]
                # -----dp结束 得到对应的数字 之后根据数字来确定哪个老师
                group_ans = []

                for num in ans:
                    temp_index = value.index(num)
                    value[temp_index] = 111111
                    group_ans.append(wait_selec_set_to_list[temp_index])

                # 确定老师 此时group_ans 就是这组迭代后最后确定的老师

                # 如果抽出来异常 打印！！！

                # ==============这段代码就是把迭代后的结果按照解的形式放到lz中取
                new_group0 = {}
                new_group1 = []

                for tea in group_ans:
                    new_group0[tea] = len(teacher_statu[tea])
                    new_group1.extend(teacher_statu[tea])

                dabian_teacher = list(group_ans)
                for stu in new_group1:
                    if id_teacher[stu] in dabian_teacher:
                        dabian_teacher.remove(id_teacher[stu])

                t_d = 0
                while (True):
                    if len(dabian_teacher) <= teachers:
                        break
                    dabian_teacher.pop(t_d)

                new_group0["teachers"] = dabian_teacher
                # =======================结束这组格式化解

                lzs[lz][n_group][0] = new_group0
                lzs[lz][n_group][1] = new_group1

                group_stu_sum = 0

                for tea in group_ans:
                    group_stu_sum += len(teacher_statu[tea])

                # 确定了一组老师 所以用本轮可用老师集合 - 用到这组老师集合 之后的几组都需要与teacher取交集
                teacher_set = teacher_set - set(group_ans)

            # 最后一组的解的格式化================
            group_stu_sum = 0
            for tea in list(teacher_set):
                group_stu_sum += len(teacher_statu[tea])

            new_group0 = {}
            new_group1 = []

            for tea in list(teacher_set):
                new_group0[tea] = len(teacher_statu[tea])
                new_group1.extend(teacher_statu[tea])

            dabian_teacher = list(teacher_set)
            for teacher in new_group1:
                if id_teacher[teacher] in dabian_teacher:
                    dabian_teacher.remove(id_teacher[teacher])

            t_d = 0
            while (True):
                if len(dabian_teacher) <= teachers:
                    break
                dabian_teacher.pop(t_d)

            new_group0["teachers"] = dabian_teacher

            lzs[lz][-1][0] = new_group0
            lzs[lz][-1][1] = new_group1
            # 格式化完成 ==================
            for group in lzs[lz]:            # 2019/10/23一起答辩
                for s_t in same_teacher:
                    res = list(set(s_t).intersection(set(list(group[0].keys())[:-1])))
                    if len(res) > 0:
                        for st in s_t:
                            if st not in group[0]['teachers']:
                                group[0]['teachers'].append(st)
                                d_t = 0
                                while (True):
                                    if group[0]['teachers'][d_t] in same_teacher:
                                        d_t += 1
                                        continue
                                    if len(group[0]['teachers']) <= teachers:
                                        break
                                    group[0]['teachers'].pop(d_t)
            for group in lzs[lz]:  # 2019/10/23不一起答辩
                for c_t in clash_teacher:
                    for ct in c_t:
                        if ct in list(group[0].keys())[:-1] and ct not in group[0]['teachers']:
                            group[0]['teachers'].append(ct)
                            d_t = 0
                            while (True):
                                if group[0]['teachers'][d_t] in c_t:
                                    d_t += 1
                                    continue
                                if len(group[0]['teachers']) <= teachers:
                                    break
                                group[0]['teachers'].pop(d_t)

            # 计算迭代后粒子的cost
            cost_2 = fit_fun.fit_all(lzs[lz], id_teacher, id_score, score_scale, n, same_teacher)
            # print("new",cost_2)
            if(cost_2 > fit[lz]):
                advance += 1
                fit[lz] = deepcopy(cost_2)
                # print('change')
            else:
                lzs[lz] = deepcopy(lz_re)

            if(cost_2 > fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher)):
                g_advance += 1
                best_indx = lz
                g_best = deepcopy(lzs[lz])
            #     print('change_g',best_indx)
            # print('\n')

        for follow in range(50):     # 每轮迭代更新50个解
            teacher_set = set(id_teacher.values())
            lz = random_pick(probability)
            lz_re = deepcopy(lzs[lz])
            # print("select",lz,fit[lz])
            for n_group in range(len(lzs[lz]) - 1):

                lz_tea_set = set(list(lzs[lz][n_group][0])[:-1])
                # 从当前粒子这组选一部分老师集合
                lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * save)))

                # 迭代后这组老师应该从这个集合中抽
                wait_selec_set_to_list = list(lz_tea_set)

                # 要与剩下老师集合做交集
                wait_selec_set_to_list = list(set(wait_selec_set_to_list) & teacher_set)

                random_tea = random.sample(list(teacher_set.difference(wait_selec_set_to_list)),
                                           round(len(lz_tea_set) * 0.5))
                wait_selec_set_to_list += random_tea

                # --dp  从n个数中抽 出x个数 使其和最接近m  关于第5点的启发式
                # 这里的n 是wait_selec_set_to_list 老师对应的学生数 m是平均每组学生数
                state = []
                for i in range(100):
                    state.append([])
                for i in state:
                    for j in range(50):
                        i.append(0)
                dp = []
                for i in range(100):
                    dp.append(0)

                value = []
                for tea in wait_selec_set_to_list:
                    value.append(len(teacher_statu[tea]))
                n2 = len(value)

                ave_stu = len(id_teacher) / len(g_best)

                sum_temp = 0
                m = round(ave_stu - sum_temp)

                for i in range(n2):
                    j = m
                    while j >= value[i]:
                        tmp = dp[j - value[i]] + value[i]
                        if tmp > dp[j]:
                            dp[j] = tmp
                            try:  ###
                                state[i][j] = 1
                            except:
                                # print("出现异常，分组过少")
                                return -1
                        j -= 1

                i = n2
                j = m
                ans = []

                while i > 0:
                    i -= 1
                    if state[i][j] == 1:
                        ans.append(value[i])
                        j -= value[i]
                # -----dp结束 得到对应的数字 之后根据数字来确定哪个老师
                group_ans = []

                for num in ans:
                    temp_index = value.index(num)
                    value[temp_index] = 111111
                    group_ans.append(wait_selec_set_to_list[temp_index])

                # 确定老师 此时group_ans 就是这组迭代后最后确定的老师

                # 如果抽出来异常 打印！！！

                # ==============这段代码就是把迭代后的结果按照解的形式放到lz中取
                new_group0 = {}
                new_group1 = []

                for tea in group_ans:
                    new_group0[tea] = len(teacher_statu[tea])
                    new_group1.extend(teacher_statu[tea])

                dabian_teacher = list(group_ans)
                for stu in new_group1:
                    if id_teacher[stu] in dabian_teacher:
                        dabian_teacher.remove(id_teacher[stu])

                t_d = 0
                while (True):
                    if len(dabian_teacher) <= teachers:
                        break
                    dabian_teacher.pop(t_d)

                new_group0["teachers"] = dabian_teacher
                # =======================结束这组格式化解

                lzs[lz][n_group][0] = new_group0
                lzs[lz][n_group][1] = new_group1

                group_stu_sum = 0

                for tea in group_ans:
                    group_stu_sum += len(teacher_statu[tea])

                # 确定了一组老师 所以用本轮可用老师集合 - 用到这组老师集合 之后的几组都需要与teacher取交集
                teacher_set = teacher_set - set(group_ans)

            # 最后一组的解的格式化================
            group_stu_sum = 0
            for tea in list(teacher_set):
                group_stu_sum += len(teacher_statu[tea])

            new_group0 = {}
            new_group1 = []

            for tea in list(teacher_set):
                new_group0[tea] = len(teacher_statu[tea])
                new_group1.extend(teacher_statu[tea])

            dabian_teacher = list(teacher_set)
            for teacher in new_group1:
                if id_teacher[teacher] in dabian_teacher:
                    dabian_teacher.remove(id_teacher[teacher])

            t_d = 0
            while (True):
                if len(dabian_teacher) <= teachers:
                    break
                dabian_teacher.pop(t_d)

            new_group0["teachers"] = dabian_teacher

            lzs[lz][-1][0] = new_group0
            lzs[lz][-1][1] = new_group1
            # 格式化完成 ==================
            for group in lzs[lz]:  # 2019/10/23一起答辩
                for s_t in same_teacher:
                    res = list(set(s_t).intersection(set(list(group[0].keys())[:-1])))
                    if len(res) > 0:
                        for st in s_t:
                            if st not in group[0]['teachers']:
                                group[0]['teachers'].append(st)
                                d_t = 0
                                while (True):
                                    if group[0]['teachers'][d_t] in same_teacher:
                                        d_t += 1
                                        continue
                                    if len(group[0]['teachers']) <= teachers:
                                        break
                                    group[0]['teachers'].pop(d_t)
            for group in lzs[lz]:  # 2019/10/23不一起答辩
                for c_t in clash_teacher:
                    for ct in c_t:
                        if ct in list(group[0].keys())[:-1] and ct not in group[0]['teachers']:
                            group[0]['teachers'].append(ct)
                            d_t = 0
                            while (True):
                                if group[0]['teachers'][d_t] in c_t:
                                    d_t += 1
                                    continue
                                if len(group[0]['teachers']) <= teachers:
                                    break
                                group[0]['teachers'].pop(d_t)

            # 计算迭代后粒子的cost
            cost_2 = fit_fun.fit_all(lzs[lz], id_teacher, id_score, score_scale, n, same_teacher)
            # print("new",cost_2)
            if (cost_2 > fit[lz]):
                advance += 1
                fit[lz] = cost_2
                # print('change')
            else:
                change_count[lz] += 1
                lzs[lz] = deepcopy(lz_re)
                if change_count[lz] > 50:
                    while True:
                        temp_lz, t_s = res_initial.lzcsh(data, n, x, 1, teachers, clash_teacher, same_teacher)
                        if fit_fun.fit_all(temp_lz[0], id_teacher, id_score, score_scale, n, same_teacher) > 0:
                            break
                    lzs[lz] = deepcopy(temp_lz[0])
                    fit[lz] = fit_fun.fit_all(temp_lz[0], id_teacher, id_score, score_scale, n, same_teacher)
                    change_count[lz] = 0

            if (cost_2 > fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher)):
                g_advance += 1
                best_indx = lz
                g_best = deepcopy(lzs[lz])
        g_best_list.append(fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher))
        # print('\n')
        # Diversity_list.append(Diversity(lzs, n))

# -----------------

    # 答辩老师数
    # f = open(r'C:\Users\44540\Desktop\czh.txt', "a")
    # ------------------------------------------------------
    # 开始迭代150轮


    # 挑出没有作为答辩老师的老师
    rest_teachers = []
    for group in g_best[:-1]:
        for i in list(group[0].keys())[:-1]:
            if i not in group[0]['teachers']:
                rest_teachers.append(i)
    # print(rest_teachers)
    # print('advance:',advance)
    # print('g_advance:',g_advance)
    # print('pre best:', pre_best)
    print('next best:', fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher),best_indx)
    # 取没有当人答辩老师的老师来补答辩老师不足的组
    for group in g_best:
        while len(group[0]['teachers']) < teachers:
            random_num = random.randint(0, len(rest_teachers) - 1)
            if rest_teachers[random_num] not in list(group[0].keys())[:-1] and teacher_statu[rest_teachers[random_num]] \
                    not in group[1] and res_initial.teacher_teacher[rest_teachers[random_num]] not in group[0][
                'teachers']:
                group[0]['teachers'].append(rest_teachers[random_num])
                rest_teachers.pop(random_num)

    # 2019/10/24
    if fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher) < 0:
        print("过多老师在同组答辩")
        return -3

    # # 打印图表
    # plot_scale(begin_best, 0,id_score)
    # plot_scale(g_best, 1,id_score)
    # plt.figure()
    # #print(len(g_best_list))

    # print(g_best_list)    # 粒子适应度
    # plt.plot(g_best_list)
    # plt.xlabel("iter_num")
    # plt.ylabel("cost (-fit)")
    # plt.show()
    #
    # print("粒子差异性",Diversity_list)     # 粒子差异性
    # plt.plot(Diversity_list)
    # plt.xlabel("iter_num")
    # plt.ylabel("Diversity_list")
    # plt.show()

    web_ans = {}
    # web_ans["status"] = 200
    for index, i in enumerate(g_best):
        web_ans[index] = {}
        web_ans[index]["dabian_teachers"] = g_best[index][0]["teachers"]
        web_ans[index]["teachers"] = {}
        sum = 0
        for teacher, stu_num in tuple(g_best[index][0].items())[:-1]:
            web_ans[index]["teachers"][teacher] = g_best[index][1][sum:sum + stu_num]
            sum += stu_num

    # 打印图表
    # plot_scale(begin_best, 0,id_score)
    plot_scale(g_best, "result", id_score)

    plt.figure()
    # #print(len(g_best_list))
    # #print(g_best_list)

    # plt.xlabel(u"迭代次数",fontproperties=font_set)
    # plt.ylabel(u"适应度",fontproperties=font_set)
    # plt.ylim(0.8,1)
    # plt.yticks(0.9,1,0.01)
    # print("===="*20)
    # print(g_best_list)
    # print("===="*20)
    # print(g_best_list[-1])



    # plt.savefig("../static/image/iterator.jpg")
    plt.savefig("static/image/iterator_07.jpg")
    plt.savefig("static/image/iterator_07.svg",format="svg")
    plt.show()

    # f.write(str(fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher)) + ' ' + str(iter_num) + "\n")
    # f.close()
    # jg用来记录txt文件，pd_data用于导入excl文件
    # jg = open(r'C:\Users\44540\Desktop\结果.txt', "w+")
    jg = open(r'files\结果(json).txt', "w+")

    #print(g_best)
    pd_data = {'学号': [], '绩点': [], '指导老师': [], '评阅老师': [], '答辩老师': []}
    for i in g_best:
        jg.write(str('{'))
        p_l = list(i[0].keys())

        for j in p_l[:-1]:
            jg.write(str('"'+j+'":'+str(i[0][j])+','))
        jg.write(str('"dabian_Teachers":['))
        for j in i[0]['teachers']:
            if i[0]['teachers'].index(j) != 0:
                jg.write(str(','))
            jg.write(str('"' + j + '"'))
        jg.write(str('],'))
        jg.write(str('"Student":['))
        for student in i[1]:
            if i[1].index(student) != 0:
                jg.write(str(','))
            jg.write(str(student))
            pd_data['学号'].append(student)
            pd_data['绩点'].append(id_score[student])
            pd_data['指导老师'].append(id_teacher[student])
            pd_data['评阅老师'].append(res_initial.student_pingyue[student])
            temp = ""
            for te in i[0]['teachers']:
                temp = temp + te + ",\n"
            pd_data['答辩老师'].append(temp)
        pd_data['学号'].append(' ')
        pd_data['绩点'].append(' ')
        pd_data['指导老师'].append(' ')
        pd_data['评阅老师'].append(' ')
        pd_data['答辩老师'].append(' ')
        jg.write(']}\n')
    jg.close()
    pd_df = pd.DataFrame(pd_data)
    pd_df.to_excel(r'C:\Users\44540\Desktop\结果表格.xlsx')
    # pd_df.to_excel(r'files\结果表格.xlsx')


    # 合并答辩老师
    head = 2
    end = 0
    # wb = load_workbook(r'C:\Users\44540\Desktop\结果表格.xlsx')
    wb = load_workbook(r'files\结果表格.xlsx')

    ws = wb.active
    for i in g_best:
        end = head + len(i[1]) - 1
        ws.merge_cells(start_row=head, start_column=6, end_row=end, end_column=6)
        ws.merge_cells(start_row=head, start_column=1, end_row=end, end_column=1)
        tem = 'A' + str(head)
        ws[tem].alignment = Alignment(horizontal='justify', vertical='center')
        ws[tem] = "第" + str(g_best.index(i)) + "组：\n" + str(len(i[1])) +"人"
        tem = 'A' + str(end+1)
        ws[tem] = " "
        tem = 'F'+str(head)
        ws[tem].alignment = Alignment(horizontal='justify', vertical='center')
        head = end + 2
    # wb.save(r'C:\Users\44540\Desktop\结果表格.xlsx')
    wb.save(r'files\结果表格.xlsx')
    #
    # web_ans["id_teacher"] = id_teacher
    # web_ans["id_score"] = id_score
    # web_ans = json.dumps(web_ans, ensure_ascii=False)
    # ##print(web_ans)
    for i in g_best:
        print(i)
    return web_ans,g_best_list


# hundred_time = []
# time = 0
# average_fit = []
# s = 0.5
# for f in range(11):
#     sum = 0
#     time = 0
#     while (time < 5):
#         try:
#             a, b = begin(n=4, accuracy_level=2, teachers=4, save=s)
#             hundred_time.append(b)
#             print("长度：", len(b))
#             time += 1
#             print("---", time, "---")
#             sum += b[-1]
#         except Exception as e:
#             pass
#         finally:
#             pass
#     print(hundred_time)
#     sum = sum/5
#     print(f,'的平均适应度',sum)
#     average_fit.append(sum)
#     s += 0.05
# plt.plot(average_fit)
# plt.xlabel("N_save")
# plt.ylabel("fitness")
# plt.xticks([0,1,2,3,4,5,6,7,8,9,10],[0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1])
# plt.show()
# for e in final:
#     print(final[e])

# print(begin(n=33, accuracy_level=1))

hundred_time = []
time = 0
while(time<1):
    # try:
        a,b = begin(n=4, accuracy_level=2, teachers=4)
        hundred_time.append(b[-1])
        print("长度：",b[-1])
        time+=1
        print("---",time,"---")
    # except Exception as e:
    #     pass
    # finally:
    #     pass
print(hundred_time)