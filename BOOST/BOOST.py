import pandas as pd
# import tools.res_initial as res_initial
from matplotlib.font_manager import FontProperties
import pandas as pd
from tools import res_initial
import json
font_set = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=12)
import math
from tools import resultToCSV
import copy
from tools import fit_fun
# import tools.fit_fun as fit_fun
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import numpy as np
import tools.plot_fun
import tools.resultToCSV
import tools.Diversity
import tools.random_pick

def if_union(a=[],b=[]):
    if len(list(set(a)&set(b))) == 0:
        return False
    else:
        return True

def show_detail(res, id_teacher):
    '''
    :param res: 初始化的多个粒子
    :return: 以文字展示
    '''
    valid_count = 0
    valid_lz = []
    for r in res:
        flag = 0
        for j, k in enumerate(r):
            flg = 0
            for te in k[0].keys():
                for su in k[1]:
                    if (te == id_teacher[su]):
                        flg = 1
                        flag = 1

        if flag == 0:
            valid_count += 1
            valid_lz.append(r)




# def plt_show_all(res, fit_list, n,id_score):
#     '''
#     :param res: res 为初始化的多个粒子
#     :param fit: 为res中粒子的适应度
#     :param n: 展示n个粒子
#     :return:
#     '''
#
#     for i, r in enumerate(list(np.argsort(np.array(fit_list[:n])))):
#         plot_scale(res[r],i,id_score)


def begin(n=4, x=2, n_groups=50, teachers=4, accuracy_level=2, clash_teacher=[], same_teacher=[], rd=0.5, alpha=0.9, save = 0.8,no_dabian=[]):
    '''
        data '评阅数据集2.xlsx'为数据
        n为组数
        x为每组人数不同的程度
        n_groups为粒子数
        teachers 为答辩老师
        rd为GA变异概率
        alpha为DE教师抽出概率
        save 为ABC原蜜源老师抽取概率
    '''
    # data = pd.read_excel(io='../input_data/new_data.xlsx')
    # data = pd.read_excel(io=r'../input_data/file.xlsx')
    data = pd.read_excel(io=r'../input_data/2020_data.xlsx',dtype=object)
    data22 = pd.read_excel(io=r'../input_data/2020_data.xlsx')
    data22.columns = ["id", "score", "teacher", "name", "theme"]

    data.columns = ["id", "score", "teacher","name","theme"]
    # 学号（学生）对应的老师
    id_teacher = dict([*zip(data["id"], data["teacher"])])
    print(data["id"])
    # （学号）学生对应的绩点
    id_score = dict([*zip(data["id"], data["score"])])
    #学号对应姓名
    id_name = dict([*zip(data["id"], data["name"])])
    # 学号对应论文题目
    id_theme = dict([*zip(data["id"], data["theme"])])

    print(id_name)
    print(id_theme)

    for k,v in id_score.items():
        id_score[k] = float(v)
    score_scale = [0, 0, 0, 0, 0]
    for k,i in id_score.items():
        if i < 2.0:
            score_scale[0] += 1
        elif i < 2.5:
            score_scale[1] += 1
        elif i < 3.0:
            score_scale[2] += 1
        elif i < 3.5:
            score_scale[3] += 1
        else:
            score_scale[4] += 1
    # 以上是对id_teacher，id_score，score_scale的初始化
    for i in range(5):
        score_scale[i] = score_scale[i] / data.shape[0]
    # 初始化函数 参数具体含义在fit

    score_series = pd.Series(data22["score"])
    # print(score_series)
    axe = plt.subplot()
    axe.hist(score_series, bins=[1, 1.5, 2, 2.5, 3, 3.5, 4],density=True)
    axe.set_xlabel(u"绩点区间",fontproperties=font_set)
    axe.set_ylabel(u"频率/间距",fontproperties=font_set)
    # axe.set_title(u"全局绩点分布",fontproperties=font_set)
    # axe.hist(score_series, bins=[0, 2, 2.5, 3, 3.5, 4],density=True)
    # plt.savefig("static/image/new_data_global.jpg")
    plt.savefig("figure/global_distribution.svg",format="svg")
    plt.savefig("figure/global_distribution.jpg")

    temp_lzcsh = res_initial.lzcsh(data, n, x, n_groups, teachers, clash_teacher, same_teacher)

    # print(temp_lzcsh)
    if isinstance(temp_lzcsh,int):
        # print("无法初始化")
        return -2
    else:
        lzs, teacher_statu = temp_lzcsh
    GA_MU_lzs = copy.deepcopy(lzs)
    DE_lzs = copy.deepcopy(lzs)
    ABC_lzs = copy.deepcopy(lzs)

    diversity_list = []
    diversity_list.append(tools.Diversity.Diversity(lzs))
    # print("初始化差异度：",diversity_list[-1])

    # res_his 记录历史最优
    # 初始化完的粒子是第一轮粒子的历史最优
    res_his = copy.deepcopy(lzs)
    cost = 0
    best_indx = 0
    fit_sum = 0     ### ABC适应度总和
    fit = []        ### ABC适应度列表
    probability = []    ### ABC蜜源抽取概率
    change_count = []   ### ABC粒子更新次数
    test_i = 0
    for lz in lzs:
        change_count.append(0)
        temp_cost = fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher)
        if temp_cost > cost:
            cost = temp_cost
            best_indx = lzs.index(lz)
        if temp_cost < 0:   ### ABC删除适应度小于0的蜜源
            test_i += 1
            while True:
                temp_lzcsh, t_s = res_initial.lzcsh(data, n, x, 1, teachers, clash_teacher, same_teacher)
                temp_cost = fit_fun.fit_all(temp_lzcsh[0], id_teacher, id_score, score_scale, n, same_teacher)
                if temp_cost > 0:
                    temp_index = lzs.index(lz)
                    lzs[temp_index] = copy.deepcopy(temp_lzcsh[0])
                    if temp_cost > cost:
                        cost = temp_cost
                        best_indx = temp_index
                    break
        # print('cccccccccccccccccc')
        fit.append(temp_cost)
        fit_sum += temp_cost

    for lz_fit in fit:
        probability.append(lz_fit/fit_sum)

    # boost_best 全局最优粒子
    boost_best =  copy.deepcopy(lzs[best_indx])
    # GA全局最优粒子 = boost_best
    g_best = boost_best
    #DE 全局最优粒子 d_best = boost_best
    d_best = boost_best
    ### ABC 全局最优粒子 a_best = boost_best
    a_best = boost_best


    # boost全局最优历史列表
    boost_best_list = [0]
    # GA全局最优历史列表
    g_best_list = [0]


    # 初始化最优粒子
    begin_best = copy.deepcopy(boost_best)

    # 迭代中出现的最优粒子
    ans_best = copy.deepcopy(g_best)
    ans_best_index = best_indx

    # 答辩老师数
    # f = open(r'C:\Users\44540\Desktop\czh.txt', "a")
    # ------------------------------------------------------
    # 开始迭代150轮
    variation_num = 0
    variation = 0
    iteration = 0
    # f = open(r'C:\Users\44540\Desktop\czh.txt', "a")


    for iter_num in range(150):
        print("当前迭代到第{}轮".format(iter_num))
        # 开始GA
        # =======================================================================================================
        # =========================================================================================================
        # =======================================================================================================
        # =======================================================================================================

        # print("开始GGGGGGGGGGGGGAAAAAAAAAAAAA")

        # 每轮进步的粒子数
        increase_sum = []

        # 每轮迭代违反硬要求的粒子数
        abnormal_count = []

        # 每轮迭代后的最好粒子
        one_iterator_best_cost = 0
        one_iterator_best_index = 6666
        # //预回滚
        lzs_re = copy.deepcopy(GA_MU_lzs)

        for index, lz in enumerate(GA_MU_lzs):
            p = 0

            # 老师集合
            teacher_set = set(id_teacher.values())

            # 计算本粒子迭代前的cost
            cost_1 = fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher)

            # 组装前n-1组 最后一组自动分出来
            for n_group in range(len(lz) - 1):
                ran_num = random.random()

                g_group_tea_set = set(list(g_best[n_group][0])[:-1])

                # 从全局最优选一部分老师集合
                g_group_tea_set = set(random.sample(g_group_tea_set, round(len(g_group_tea_set) * 0.7)))

                p_group_tea_set = set(list(res_his[index][n_group][0])[:-1])
                # 从历史最优选一部分老师集合
                p_group_tea_set = set(random.sample(p_group_tea_set, round(len(p_group_tea_set) * 0.7)))

                lz_tea_set = set(list(lz[n_group][0])[:-1])
                # 从当前粒子这组选一部分老师集合
                lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * 0.5)))

                # 迭代后这组老师应该从这个集合中抽
                wait_selec_set_to_list = list(g_group_tea_set | p_group_tea_set | lz_tea_set)

                # 要与剩下老师集合做交集
                wait_selec_set_to_list = list(set(wait_selec_set_to_list) & teacher_set)

                # if ran_num < 0.1:
                if ran_num < rd:
                    #print("*************发送变异************")
                    select_num = 3
                    while select_num > 0:
                        if select_num < len(list(teacher_set.difference(wait_selec_set_to_list))):
                            random_tea = random.sample(list(teacher_set.difference(wait_selec_set_to_list)), select_num)
                            break
                        else:
                            select_num -= 1
                    if select_num != 0:
                        wait_selec_set_to_list += random_tea

                # --dp  从n个数中抽 出x个数 使其和最接近m  关于第4点的启发式
                # 这里的n 是wait_selec_set_to_list 老师对应的学生数 m是平均每组学生数
                state = []
                for i in range(100):
                    state.append([])
                for i in state:
                    for j in range(50):
                        i.append(0)
                dp = []
                for i in range(100):
                    dp.append(0)

                value = []
                for tea in wait_selec_set_to_list:
                    value.append(len(teacher_statu[tea]))
                n_dp = len(value)

                ave_stu = len(id_teacher) / len(g_best)

                sum_temp = 0
                # m = random.randint(round(ave_stu * 0.9),round(ave_stu * 1.1))
                # m = random.randint(round(ave_stu-2),round(ave_stu+2))
                m = round(ave_stu)
                # print("Jun",m)
                # m = round()

                for i in range(n_dp):
                    j = m
                    while j >= value[i]:
                        tmp = dp[j - value[i]] + value[i]
                        if tmp > dp[j]:
                            dp[j] = tmp
                            try:                        ###
                                state[i][j] = 1
                            except:
                                #print("出现异常，分组过少")
                                return -1
                        j -= 1

                i = n_dp
                j = m
                ans = []

                while i > 0:
                    i -= 1
                    if state[i][j] == 1:
                        ans.append(value[i])
                        j -= value[i]
                # -----dp结束 得到对应的数字 之后根据数字来确定哪个老师
                group_ans = []

                for num in ans:
                    temp_index = value.index(num)
                    value[temp_index] = 111111
                    group_ans.append(wait_selec_set_to_list[temp_index])

                # 确定老师 此时group_ans 就是这组迭代后最后确定的老师

                # 如果抽出来异常 打印！！！

                # if ran_num < 0.1:
                if ran_num < rd:
                    if len(set(random_tea).union(set(group_ans))) != 0:
                        variation += 1
                        p = 1

                # ==============这段代码就是把迭代后的结果按照解的形式放到lz中取
                new_group0 = {}
                new_group1 = []

                for tea in group_ans:
                    new_group0[tea] = len(teacher_statu[tea])
                    new_group1.extend(teacher_statu[tea])

                dabian_teacher = list(group_ans)
                for stu in new_group1:
                    if id_teacher[stu] in dabian_teacher:
                        dabian_teacher.remove(id_teacher[stu])

                t_d = 0
                while (True):
                    if len(dabian_teacher) <= teachers:
                        break
                    dabian_teacher.pop(t_d)

                new_group0["teachers"] = dabian_teacher
                # =======================结束这组格式化解

                lz[n_group][0] = new_group0
                lz[n_group][1] = new_group1

                group_stu_sum = 0

                for tea in group_ans:
                    group_stu_sum += len(teacher_statu[tea])

                # 确定了一组老师 所以用本轮可用老师集合 - 用到这组老师集合 之后的几组都需要与teacher取交集
                teacher_set = teacher_set - set(group_ans)

            # 最后一组的解的格式化================
            group_stu_sum = 0
            for tea in list(teacher_set):
                group_stu_sum += len(teacher_statu[tea])

            new_group0 = {}
            new_group1 = []

            for tea in list(teacher_set):
                new_group0[tea] = len(teacher_statu[tea])
                new_group1.extend(teacher_statu[tea])

            dabian_teacher = list(teacher_set)
            for teacher in new_group1:
                if id_teacher[teacher] in dabian_teacher:
                    dabian_teacher.remove(id_teacher[teacher])

            t_d = 0
            while (True):
                if len(dabian_teacher) <= teachers:
                    break
                dabian_teacher.pop(t_d)

            new_group0["teachers"] = dabian_teacher

            lz[-1][0] = new_group0
            lz[-1][1] = new_group1

            # print("xxx"*20)
            # for i in lz:
            #     print(len(i[-1]))

            # 格式化完成 ==================
            for group in lz:            # 2019/10/23一起答辩
                for s_t in same_teacher:
                    res = list(set(s_t).intersection(set(list(group[0].keys())[:-1])))
                    if len(res) > 0:
                        for st in s_t:
                            if st not in group[0]['teachers']:
                                group[0]['teachers'].append(st)
                                d_t = 0
                                while (True):
                                    if group[0]['teachers'][d_t] in same_teacher:
                                        d_t += 1
                                        continue
                                    if len(group[0]['teachers']) <= teachers:
                                        break
                                    group[0]['teachers'].pop(d_t)
            for group in lz:  # 2019/10/23不一起答辩
                for c_t in clash_teacher:
                    for ct in c_t:
                        if ct in group[0]['teachers']:
                            d_t = 0
                            while True:
                                if group[0]['teachers'][d_t] != ct and group[0]['teachers'][d_t] in c_t:
                                    group[0]['teachers'].pop(d_t)
                                d_t += 1
                                if d_t >= len(group[0]['teachers']):
                                    break
            for group in lz:
                print(group)
            print('/n')
            # 计算迭代后粒子的cost
            cost_2 = fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher)

            # 和历史最优比较
            if cost_2 > fit_fun.fit_all(res_his[index], id_teacher, id_score, score_scale, n, same_teacher):

                res_his[index] = copy.deepcopy(lz)

            # 记录所有迭代中出现的最好粒子
            if cost_2 > fit_fun.fit_all(ans_best, id_teacher, id_score, score_scale, n, same_teacher):
                ans_best = copy.deepcopy(lz)

            # 和全局最优进行比较
            if cost_2 > fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher):
                iteration = -1
                print("GA 全局最优发生变化:",cost_2)
                g_best = copy.deepcopy(lz)


                a,b = tools.fit_fun.show_fit(g_best, id_teacher, id_score, score_scale, n, same_teacher)
                print("fit:")
                print(a+b)
            # 计算本轮出现最好粒子
            if cost_2 > one_iterator_best_cost:
                one_iterator_best_cost = cost_2
                one_iterator_best_index = index

            # 异常
            if cost_2 < 0.0001:
                abnormal_count.append(index)
            # 进步
            if cost_2 > cost_1:
                # #print(index)

                if p == 1:
                    variation_num += 1

                increase_sum.append(index)
        if one_iterator_best_cost < g_best_list[-1]:
            GA_MU_lzs = copy.deepcopy(lzs_re)
            g_best_list.append(g_best_list[-1])
        # g_best = copy.deepcopy(lzs[index])
        else:
            g_best_list.append(one_iterator_best_cost)


        g_diversity = tools.Diversity.Diversity(GA_MU_lzs)
        # print("第{}轮迭代差异度：".format(iter_num),diversity_list[-1])

    # 开始DE
    # =======================================================================================================
    # =========================================================================================================
    # =======================================================================================================
    # =======================================================================================================
    #     print("开始DDDDDDDDDDDDDDDDDDDEEEEEEEEEEEEEEEEE")
        for index, lz in enumerate(DE_lzs):


            # print("当前粒子")
            # for  i in lz:
            #     print(len(i[-1]))
            # if index == 3:
            #     print("xxxx"*20)
            #     print(1)
            #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))


            p = 0
            H = []
            for i in range(len(lz)):
                H.append([])
            V = []
            for i in range(len(lz)):
                V.append([])
            # 老师集合
            teacher_set = set(id_teacher.values())

            # 计算本粒子迭代前的cost
            cost_1 = fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher)

            three_sample = random.sample(range(0, len(DE_lzs) - 1), 3)
            p = []
            for i in range(3):
                p.append(copy.deepcopy(DE_lzs[three_sample[i]]))
            fit_p1 = fit_fun.fit_all(p[1], id_teacher, id_score, score_scale, n, same_teacher)
            fit_p2 = fit_fun.fit_all(p[2], id_teacher, id_score, score_scale, n, same_teacher)
            if (fit_p2 > fit_p1):
                p[1], p[2] = p[2], p[1]

            p[1] = d_best

            # if index == 3:
            #     print(2)
            #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))

            # 组装第一个中间变量H的前n-1组 最后一组自动分出来
            for n_group in range(len(lz) - 1):

                # ran_num = random.random()
                #
                # g_group_tea_set = set(list(g_best[n_group][0])[:-1])
                #
                # # 从全局最优选一部分老师集合
                # g_group_tea_set = set(random.sample(g_group_tea_set, round(len(g_group_tea_set) * 0.7)))
                #
                # p_group_tea_set = set(list(res_his[index][n_group][0])[:-1])
                # # 从历史最优选一部分老师集合
                # p_group_tea_set = set(random.sample(p_group_tea_set, round(len(p_group_tea_set) * 0.7)))
                #
                # lz_tea_set = set(list(lz[n_group][0])[:-1])
                # # 从当前粒子这组选一部分老师集合
                # lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * 0.5)))



                p1_ReviewT = set(list(p[1][n_group][0])[:-1])
                p2_ReviewT = set(list(p[2][n_group][0])[:-1])
                p0_ReviewT = set(list(p[0][n_group][0])[:-1])


                # print(p1_ReviewT)
                # print(p1_ReviewT-p2_ReviewT)
                # if index == 3:
                #     print(2.1)
                #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))



                # 迭代后这组老师应该从这个集合中抽
                wait_selec_set_to_list = list(p0_ReviewT | (p1_ReviewT - p2_ReviewT))

                # 要与剩下老师集合做交集
                wait_selec_set_to_list = list(set(wait_selec_set_to_list) & teacher_set)

                # if ran_num < 0.1:
                #     #print("*************发送变异************")
                #     random_tea = random.sample(list(teacher_set.difference(wait_selec_set_to_list)), 3)
                #     wait_selec_set_to_list += random_tea

                # --dp  从n个数中抽 出x个数 使其和最接近m  关于第5点的启发式
                # 这里的n 是wait_selec_set_to_list 老师对应的学生数 m是平均每组学生数
                state = []
                for i in range(100):
                    state.append([])
                for i in state:
                    for j in range(50):
                        i.append(0)
                dp = []
                for i in range(100):
                    dp.append(0)

                value = []
                for tea in wait_selec_set_to_list:
                    value.append(len(teacher_statu[tea]))
                n_dp = len(value)

                ave_stu = len(id_teacher) / len(d_best)

                sum_temp = 0
                m = round(ave_stu - sum_temp)

                # if index == 3:
                #     print(2.2)
                #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))

                for i in range(n_dp):
                    j = m
                    while j >= value[i]:
                        tmp = dp[j - value[i]] + value[i]
                        if tmp > dp[j]:
                            dp[j] = tmp
                            try:                        ###
                                state[i][j] = 1
                            except:
                                #print("出现异常，分组过少")
                                return -1
                        j -= 1

                i = n_dp
                j = m
                ans = []

                while i > 0:
                    i -= 1
                    if state[i][j] == 1:
                        ans.append(value[i])
                        j -= value[i]
                # -----dp结束 得到对应的数字 之后根据数字来确定哪个老师
                group_ans = []

                for num in ans:
                    temp_index = value.index(num)
                    value[temp_index] = 111111
                    group_ans.append(wait_selec_set_to_list[temp_index])

                # print("ans:",group_ans)

                # 确定老师 此时group_ans 就是这组迭代后最后确定的老师

                # 如果抽出来异常 打印！！！

                # if ran_num < 0.1:
                #     if len(set(random_tea).union(set(group_ans))) != 0:
                #         variation += 1
                #         p = 1

                # ==============这段代码就是把迭代后的结果按照解的形式放到lz中取
                new_group0 = {}
                new_group1 = []

                # if index == 3:
                #     print(2.3)
                #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))

                for tea in group_ans:
                    new_group0[tea] = len(teacher_statu[tea])
                    new_group1.extend(teacher_statu[tea])

                dabian_teacher = list(group_ans)
                for stu in new_group1:
                    if id_teacher[stu] in dabian_teacher:
                        dabian_teacher.remove(id_teacher[stu])

                t_d = 0
                while (True):
                    if len(dabian_teacher) <= teachers:
                        break
                    dabian_teacher.pop(t_d)

                new_group0["teachers"] = dabian_teacher
                # =======================结束这组格式化解



                H[n_group].append(new_group0)
                H[n_group].append(new_group1)
                # print("anssss!!!")
                # print(len(H))
                # print(H)

                # lz[n_group][0] = new_group0
                # lz[n_group][1] = new_group1

                group_stu_sum = 0

                # print("group_ans")
                # print(group_ans)
                for tea in group_ans:
                    group_stu_sum += len(teacher_statu[tea])
                # print("第{}组:".format(n_group),group_stu_sum)

                # 确定了一组老师 所以用本轮可用老师集合 - 用到这组老师集合 之后的几组都需要与teacher取交集
                teacher_set = teacher_set - set(group_ans)

            # if index == 3:
            #     print(3)
            #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))

            # 最后一组的解的格式化================
            group_stu_sum = 0
            for tea in list(teacher_set):
                group_stu_sum += len(teacher_statu[tea])

            new_group0 = {}
            new_group1 = []

            for tea in list(teacher_set):
                new_group0[tea] = len(teacher_statu[tea])
                new_group1.extend(teacher_statu[tea])

            dabian_teacher = list(teacher_set)
            for teacher in new_group1:
                if id_teacher[teacher] in dabian_teacher:
                    dabian_teacher.remove(id_teacher[teacher])

            # if index == 3:
            #     print(4)
            #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))

            t_d = 0
            while (True):
                if len(dabian_teacher) <= teachers:
                    break
                dabian_teacher.pop(t_d)

            new_group0["teachers"] = dabian_teacher

            H[-1].append(new_group0)
            H[-1].append(new_group1)

            # if index == 3:
            #     print(5)
            #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))
            # 格式化完成 ==================
            # for group in lz:            # 2019/10/23一起答辩
            #     for s_t in same_teacher:
            #         res = list(set(s_t).intersection(set(list(group[0].keys())[:-1])))
            #         if len(res) > 0:
            #             for st in s_t:
            #                 if st not in group[0]['teachers']:
            #                     group[0]['teachers'].append(st)
            #                     d_t = 0
            #                     while (True):
            #                         if group[0]['teachers'][d_t] in same_teacher:
            #                             d_t += 1
            #                             continue
            #                         if len(group[0]['teachers']) <= teachers:
            #                             break
            #                         group[0]['teachers'].pop(d_t)
            # for group in lz:  # 2019/10/23不一起答辩
            #     for c_t in clash_teacher:
            #         for ct in c_t:
            #             if ct in list(group[0].keys())[:-1] and ct not in group[0]['teachers']:
            #                 group[0]['teachers'].append(ct)
            #                 d_t = 0
            #                 while (True):
            #                     if group[0]['teachers'][d_t] in c_t:
            #                         d_t += 1
            #                         continue
            #                     if len(group[0]['teachers']) <= teachers:
            #                         break
            #                     group[0]['teachers'].pop(d_t)

            # print("第一个中间变量H")
            # print(H)
            # for i in H:
            #     print(len(i[1]))

            # 第二个中间变量V
            teacher_set = set(id_teacher.values())
            for n_group in range(len(lz) - 1):

                # ran_num = random.random()
                #
                # g_group_tea_set = set(list(g_best[n_group][0])[:-1])
                #
                # # 从全局最优选一部分老师集合
                # g_group_tea_set = set(random.sample(g_group_tea_set, round(len(g_group_tea_set) * 0.7)))
                #

                # ---使用针对每个老师以一定概率抽出
                # H_tea_set = set()
                # for teacher in list(H[n_group][0])[:-1]:
                #     if random.random()<alpha:
                #         H_tea_set.add(teacher)


                # 直接从这组老师中抽出百分之一定概率
                H_tea_set = set(list(H[n_group][0])[:-1])
                # 从历史最优选一部分老师集合
                H_tea_set = set(random.sample(H_tea_set, round(len(H_tea_set) * alpha)))
                # H_tea_set = set(random.sample(H_tea_set, round(len(H_tea_set) * alpha)))

                # ---使用针对每个老师以一定概率抽出
                # lz_tea_set = set()
                # for teacher in list(lz[n_group][0])[:-1]:
                #     if random.random() < (1-alpha):
                #         lz_tea_set.add(teacher)

                # 直接从这组老师中抽出百分之一定概率
                lz_tea_set = set(list(lz[n_group][0])[:-1])
                # 从当前粒子这组选一部分老师集合
                lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * alpha)))
                # lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * alpha)))
                # lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * (1-alpha))))


                # print(p1_ReviewT)
                # print(p1_ReviewT-p2_ReviewT)

                # 迭代后这组老师应该从这个集合中抽
                wait_selec_set_to_list = list(lz_tea_set | H_tea_set)

                # 要与剩下老师集合做交集
                wait_selec_set_to_list = list(set(wait_selec_set_to_list) & teacher_set)

                # if ran_num < 0.1:
                #     #print("*************发送变异************")
                #     random_tea = random.sample(list(teacher_set.difference(wait_selec_set_to_list)), 3)
                #     wait_selec_set_to_list += random_tea

                # --dp  从n个数中抽 出x个数 使其和最接近m  关于第5点的启发式
                # 这里的n 是wait_selec_set_to_list 老师对应的学生数 m是平均每组学生数
                state = []
                for i in range(100):
                    state.append([])
                for i in state:
                    for j in range(50):
                        i.append(0)
                dp = []
                for i in range(100):
                    dp.append(0)

                value = []
                for tea in wait_selec_set_to_list:
                    value.append(len(teacher_statu[tea]))
                n_dp = len(value)

                ave_stu = len(id_teacher) / len(g_best)

                sum_temp = 0
                m = round(ave_stu - sum_temp)

                for i in range(n_dp):
                    j = m
                    while j >= value[i]:
                        tmp = dp[j - value[i]] + value[i]
                        if tmp > dp[j]:
                            dp[j] = tmp
                            try:  ###
                                state[i][j] = 1
                            except:
                                # print("出现异常，分组过少")
                                return -1
                        j -= 1

                i = n_dp
                j = m
                ans = []

                while i > 0:
                    i -= 1
                    if state[i][j] == 1:
                        ans.append(value[i])
                        j -= value[i]
                # -----dp结束 得到对应的数字 之后根据数字来确定哪个老师
                group_ans = []

                for num in ans:
                    temp_index = value.index(num)
                    value[temp_index] = 111111
                    group_ans.append(wait_selec_set_to_list[temp_index])

                # print("ans:",group_ans)

                # 确定老师 此时group_ans 就是这组迭代后最后确定的老师

                # 如果抽出来异常 打印！！！

                # if ran_num < 0.1:
                #     if len(set(random_tea).union(set(group_ans))) != 0:
                #         variation += 1
                #         p = 1

                # ==============这段代码就是把迭代后的结果按照解的形式放到lz中取
                new_group0 = {}
                new_group1 = []

                for tea in group_ans:
                    new_group0[tea] = len(teacher_statu[tea])
                    new_group1.extend(teacher_statu[tea])

                dabian_teacher = list(group_ans)
                for stu in new_group1:
                    if id_teacher[stu] in dabian_teacher:
                        dabian_teacher.remove(id_teacher[stu])

                t_d = 0
                while (True):
                    if len(dabian_teacher) <= teachers:
                        break
                    dabian_teacher.pop(t_d)

                new_group0["teachers"] = dabian_teacher
                # =======================结束这组格式化解

                V[n_group].append(new_group0)
                V[n_group].append(new_group1)
                # print("anssss!!!")
                # print(len(H))
                # print(H)

                # lz[n_group][0] = new_group0
                # lz[n_group][1] = new_group1

                group_stu_sum = 0

                for tea in group_ans:
                    group_stu_sum += len(teacher_statu[tea])

                # 确定了一组老师 所以用本轮可用老师集合 - 用到这组老师集合 之后的几组都需要与teacher取交集
                teacher_set = teacher_set - set(group_ans)

            # 最后一组的解的格式化================
            group_stu_sum = 0
            for tea in list(teacher_set):
                group_stu_sum += len(teacher_statu[tea])

            new_group0 = {}
            new_group1 = []

            for tea in list(teacher_set):
                new_group0[tea] = len(teacher_statu[tea])
                new_group1.extend(teacher_statu[tea])

            dabian_teacher = list(teacher_set)
            for teacher in new_group1:
                if id_teacher[teacher] in dabian_teacher:
                    dabian_teacher.remove(id_teacher[teacher])

            t_d = 0
            while (True):
                if len(dabian_teacher) <= teachers:
                    break
                dabian_teacher.pop(t_d)

            new_group0["teachers"] = dabian_teacher

            V[-1].append(new_group0)
            V[-1].append(new_group1)

            # if index == 3:
            #     print(6)
            #     print(fit_fun.fit_all(lz, id_teacher, id_score, score_scale, n, same_teacher))
            # 第二个中间变量结束

            # print("----------H------------",fit_fun.fit_all(H, id_teacher, id_score, score_scale, n, same_teacher))
            # print("----------V------------",fit_fun.fit_all(V, id_teacher, id_score, score_scale, n, same_teacher))

            # 计算迭代后粒子的cost
            # print("第22222个中间变量")
            # print(V)
            # flg = 0
            for i in V:
                if len(i[1])==0: #如果某一组学生数目为0
                 V = H

            cost_v = fit_fun.fit_all(V, id_teacher, id_score, score_scale, n, same_teacher)

            # 和历史最优比较
            # if cost_2 > fit_fun.fit_all(res_his[index], id_teacher, id_score, score_scale, n, same_teacher):
            #
            #     res_his[index] = copy.deepcopy(lz)

            # 记录所有迭代中出现的最好粒子
            # if cost_2 > fit_fun.fit_all(ans_best, id_teacher, id_score, score_scale, n, same_teacher):
            #     ans_best = copy.deepcopy(lz)

            # 和全局最优进行比较
            if cost_v > fit_fun.fit_all(d_best, id_teacher, id_score, score_scale, n, same_teacher):
                iteration = -1
                print("DE 全局最优发生变化:",cost_v)
                d_best = copy.deepcopy(V)

            # 计算本轮出现最好粒子

            # if max(cost_1,cost_v) > one_iterator_best_cost:
            #     one_iterator_best_cost = max(cost_1,cost_v)
            #     one_iterator_best_index = index

            # 异常
            if cost_v < 0.0001:
                abnormal_count.append(index)
            # 进步

            # if index == 3:
            #     print("第{}轮：".format(iter_num))
            #     print(cost_1)
            #     print(cost_v)
            #     print("_________"*10)
            if cost_v > cost_1:
                # #print(index)


                # if p == 1:
                #     variation_num += 1
                DE_lzs[index] = copy.deepcopy(V)
                # increase_sum.append(index)
                # if(cost_v>one_iterator_best_cost):
                #     # print("第{}轮，")
                #     one_iterator_best_cost = cost_v
            # one_iterator_best_cost = max(one_iterator_best_cost,max(cost_v,cost_1))
# 564-611

        # if one_iterator_best_cost < g_best_list[-1]:
        #     lzs = copy.deepcopy(lzs_re)
        #     g_best_list.append(g_best_list[-1])
        # # g_best = copy.deepcopy(lzs[index])
        # else:
        #     g_best_list.append(one_iterator_best_cost)

        # print("g_best:",g_best)
        # print("one_iter_best",one_iterator_best_cost)

        d_diversity = tools.Diversity.Diversity(DE_lzs)
        # print("第{}轮迭代差异度:".format(iter_num),diversity_list[-1])

# ABC开始
# =======================================================================================================
# =======================================================================================================
# =======================================================================================================
# =======================================================================================================
        for lz in range(len(ABC_lzs)):

            # 预回滚
            lz_re = copy.deepcopy(ABC_lzs[lz])

            # 老师集合
            teacher_set = set(id_teacher.values())

            # 组装前n-1组 最后一组自动分出来
            for n_group in range(len(ABC_lzs[lz]) - 1):

                lz_tea_set = set(list(ABC_lzs[lz][n_group][0])[:-1])
                # 从当前粒子这组选一部分老师集合
                lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * save)))

                # 迭代后这组老师应该从这个集合中抽
                wait_selec_set_to_list = list(lz_tea_set)

                # 要与剩下老师集合做交集
                wait_selec_set_to_list = list(set(wait_selec_set_to_list) & teacher_set)

                # 随机更新
                random_tea = random.sample(list(teacher_set.difference(wait_selec_set_to_list)),
                                           round(len(lz_tea_set) * 0.5))
                wait_selec_set_to_list += random_tea

                # --dp  从n个数中抽 出x个数 使其和最接近m  关于第5点的启发式
                # 这里的n 是wait_selec_set_to_list 老师对应的学生数 m是平均每组学生数
                state = []
                for i in range(100):
                    state.append([])
                for i in state:
                    for j in range(50):
                        i.append(0)
                dp = []
                for i in range(100):
                    dp.append(0)

                value = []
                for tea in wait_selec_set_to_list:
                    value.append(len(teacher_statu[tea]))
                n_dp = len(value)

                ave_stu = len(id_teacher) / len(a_best)

                sum_temp = 0
                m = round(ave_stu - sum_temp)

                for i in range(n_dp):
                    j = m
                    while j >= value[i]:
                        tmp = dp[j - value[i]] + value[i]
                        if tmp > dp[j]:
                            dp[j] = tmp
                            try:  ###
                                state[i][j] = 1
                            except:
                                # print("出现异常，分组过少")
                                return -1
                        j -= 1

                i = n_dp
                j = m
                ans = []

                while i > 0:
                    i -= 1
                    if state[i][j] == 1:
                        ans.append(value[i])
                        j -= value[i]
                # -----dp结束 得到对应的数字 之后根据数字来确定哪个老师
                group_ans = []

                for num in ans:
                    temp_index = value.index(num)
                    value[temp_index] = 111111
                    group_ans.append(wait_selec_set_to_list[temp_index])

                # 确定老师 此时group_ans 就是这组迭代后最后确定的老师

                # 如果抽出来异常 打印！！！

                # ==============这段代码就是把迭代后的结果按照解的形式放到lz中取
                new_group0 = {}
                new_group1 = []

                for tea in group_ans:
                    new_group0[tea] = len(teacher_statu[tea])
                    new_group1.extend(teacher_statu[tea])

                dabian_teacher = list(group_ans)
                for stu in new_group1:
                    if id_teacher[stu] in dabian_teacher:
                        dabian_teacher.remove(id_teacher[stu])

                t_d = 0
                while (True):
                    if len(dabian_teacher) <= teachers:
                        break
                    dabian_teacher.pop(t_d)

                new_group0["teachers"] = dabian_teacher
                # =======================结束这组格式化解

                ABC_lzs[lz][n_group][0] = new_group0
                ABC_lzs[lz][n_group][1] = new_group1

                group_stu_sum = 0

                for tea in group_ans:
                    group_stu_sum += len(teacher_statu[tea])

                # 确定了一组老师 所以用本轮可用老师集合 - 用到这组老师集合 之后的几组都需要与teacher取交集
                teacher_set = teacher_set - set(group_ans)

            # 最后一组的解的格式化================
            group_stu_sum = 0
            for tea in list(teacher_set):
                group_stu_sum += len(teacher_statu[tea])

            new_group0 = {}
            new_group1 = []

            for tea in list(teacher_set):
                new_group0[tea] = len(teacher_statu[tea])
                new_group1.extend(teacher_statu[tea])

            dabian_teacher = list(teacher_set)
            for teacher in new_group1:
                if id_teacher[teacher] in dabian_teacher:
                    dabian_teacher.remove(id_teacher[teacher])

            t_d = 0
            while (True):
                if len(dabian_teacher) <= teachers:
                    break
                dabian_teacher.pop(t_d)

            new_group0["teachers"] = dabian_teacher

            ABC_lzs[lz][-1][0] = new_group0
            ABC_lzs[lz][-1][1] = new_group1
            # 格式化完成 ==================
            for group in ABC_lzs[lz]:  # 2019/10/23一起答辩
                for s_t in same_teacher:
                    res = list(set(s_t).intersection(set(list(group[0].keys())[:-1])))
                    if len(res) > 0:
                        for st in s_t:
                            if st not in group[0]['teachers']:
                                group[0]['teachers'].append(st)
                                d_t = 0
                                while (True):
                                    if group[0]['teachers'][d_t] in same_teacher:
                                        d_t += 1
                                        continue
                                    if len(group[0]['teachers']) <= teachers:
                                        break
                                    group[0]['teachers'].pop(d_t)
            for group in ABC_lzs[lz]:  # 2019/10/23不一起答辩
                for c_t in clash_teacher:
                    for ct in c_t:
                        if ct in group[0]['teachers']:
                            d_t = 0
                            while True:
                                if group[0]['teachers'][d_t] != ct and group[0]['teachers'][d_t] in c_t:
                                    group[0]['teachers'].pop(d_t)
                                d_t += 1
                                if d_t >= len(group[0]['teachers']):
                                    break

            # 计算迭代后粒子的cost
            cost_2 = fit_fun.fit_all(ABC_lzs[lz], id_teacher, id_score, score_scale, n, same_teacher)
            # 和原先比较
            if (cost_2 > fit[lz]):
                fit[lz] = copy.deepcopy(cost_2)
                # print('change')
            else:
                ABC_lzs[lz] = copy.deepcopy(lz_re)

            if (cost_2 > fit_fun.fit_all(a_best, id_teacher, id_score, score_scale, n, same_teacher)):
                a_best = copy.deepcopy(ABC_lzs[lz])
                print("ABC 全局最优发生变化",cost_2)
            #     print('change_g',best_indx)
            # print('\n')

        for follow in range(20):  # 每轮迭代更新20个解
            teacher_set = set(id_teacher.values())
            lz = tools.random_pick.random_pick(probability)
            lz_re = copy.deepcopy(ABC_lzs[lz])
            # print("select",lz,fit[lz])
            for n_group in range(len(ABC_lzs[lz]) - 1):

                lz_tea_set = set(list(ABC_lzs[lz][n_group][0])[:-1])
                # 从当前粒子这组选一部分老师集合
                lz_tea_set = set(random.sample(lz_tea_set, round(len(lz_tea_set) * save)))

                # 迭代后这组老师应该从这个集合中抽
                wait_selec_set_to_list = list(lz_tea_set)

                # 要与剩下老师集合做交集
                wait_selec_set_to_list = list(set(wait_selec_set_to_list) & teacher_set)

                random_tea = random.sample(list(teacher_set.difference(wait_selec_set_to_list)),
                                           round(len(lz_tea_set) * 0.5))
                wait_selec_set_to_list += random_tea

                # --dp  从n个数中抽 出x个数 使其和最接近m  关于第5点的启发式
                # 这里的n 是wait_selec_set_to_list 老师对应的学生数 m是平均每组学生数
                state = []
                for i in range(100):
                    state.append([])
                for i in state:
                    for j in range(50):
                        i.append(0)
                dp = []
                for i in range(100):
                    dp.append(0)

                value = []
                for tea in wait_selec_set_to_list:
                    value.append(len(teacher_statu[tea]))
                n_dp = len(value)

                ave_stu = len(id_teacher) / len(a_best)

                sum_temp = 0
                m = round(ave_stu - sum_temp)

                for i in range(n_dp):
                    j = m
                while j >= value[i]:
                    tmp = dp[j - value[i]] + value[i]
                    if tmp > dp[j]:
                        dp[j] = tmp
                        try:  ###
                            state[i][j] = 1
                        except:
                            # print("出现异常，分组过少")
                            return -1
                    j -= 1

                i = n_dp
                j = m
                ans = []

                while i > 0:
                    i -= 1
                    if state[i][j] == 1:
                        ans.append(value[i])
                        j -= value[i]
                # -----dp结束 得到对应的数字 之后根据数字来确定哪个老师
                group_ans = []

                for num in ans:
                    temp_index = value.index(num)
                    value[temp_index] = 111111
                    group_ans.append(wait_selec_set_to_list[temp_index])

                # 确定老师 此时group_ans 就是这组迭代后最后确定的老师

                # 如果抽出来异常 打印！！！

                # ==============这段代码就是把迭代后的结果按照解的形式放到lz中取
                new_group0 = {}
                new_group1 = []

                for tea in group_ans:
                    new_group0[tea] = len(teacher_statu[tea])
                    new_group1.extend(teacher_statu[tea])

                dabian_teacher = list(group_ans)
                for stu in new_group1:
                    if id_teacher[stu] in dabian_teacher:
                        dabian_teacher.remove(id_teacher[stu])

                t_d = 0
                while (True):
                    if len(dabian_teacher) <= teachers:
                        break
                    dabian_teacher.pop(t_d)

                new_group0["teachers"] = dabian_teacher
                # =======================结束这组格式化解

                ABC_lzs[lz][n_group][0] = new_group0
                ABC_lzs[lz][n_group][1] = new_group1

                group_stu_sum = 0

                for tea in group_ans:
                    group_stu_sum += len(teacher_statu[tea])

                # 确定了一组老师 所以用本轮可用老师集合 - 用到这组老师集合 之后的几组都需要与teacher取交集
                teacher_set = teacher_set - set(group_ans)

            # 最后一组的解的格式化================
            group_stu_sum = 0
            for tea in list(teacher_set):
                group_stu_sum += len(teacher_statu[tea])

            new_group0 = {}
            new_group1 = []

            for tea in list(teacher_set):
                new_group0[tea] = len(teacher_statu[tea])
                new_group1.extend(teacher_statu[tea])

            dabian_teacher = list(teacher_set)
            for teacher in new_group1:
                if id_teacher[teacher] in dabian_teacher:
                    dabian_teacher.remove(id_teacher[teacher])

            t_d = 0
            while (True):
                if len(dabian_teacher) <= teachers:
                    break
                dabian_teacher.pop(t_d)

            new_group0["teachers"] = dabian_teacher

            ABC_lzs[lz][-1][0] = new_group0
            ABC_lzs[lz][-1][1] = new_group1
            # 格式化完成 ==================
            for group in ABC_lzs[lz]:  # 2019/10/23一起答辩
                for s_t in same_teacher:
                    res = list(set(s_t).intersection(set(list(group[0].keys())[:-1])))
                    if len(res) > 0:
                        for st in s_t:
                            if st not in group[0]['teachers']:
                                group[0]['teachers'].append(st)
                                d_t = 0
                                while (True):
                                    if group[0]['teachers'][d_t] in same_teacher:
                                        d_t += 1
                                        continue
                                    if len(group[0]['teachers']) <= teachers:
                                        break
                                    group[0]['teachers'].pop(d_t)
            for group in ABC_lzs[lz]:  # 2019/10/23不一起答辩
                for c_t in clash_teacher:
                    for ct in c_t:
                        if ct in group[0]['teachers']:
                            d_t = 0
                            while True:
                                if group[0]['teachers'][d_t] != ct and group[0]['teachers'][d_t] in c_t:
                                    group[0]['teachers'].pop(d_t)
                                d_t += 1
                                if d_t >= len(group[0]['teachers']):
                                    break

            # 计算迭代后粒子的cost
            cost_2 = fit_fun.fit_all(ABC_lzs[lz], id_teacher, id_score, score_scale, n, same_teacher)
            # print("new",cost_2)
            if (cost_2 > fit[lz]):
                fit[lz] = cost_2
                # print('change')
            else:
                change_count[lz] += 1
                ABC_lzs[lz] = copy.deepcopy(lz_re)
                if change_count[lz] > 20:
                    while True:
                        temp_lz, t_s = res_initial.lzcsh(data, n, x, 1, teachers, clash_teacher, same_teacher)
                        if fit_fun.fit_all(temp_lz[0], id_teacher, id_score, score_scale, n, same_teacher) > 0:
                            ABC_lzs[lz] = copy.deepcopy(temp_lz[0])
                            break
                    fit[lz] = fit_fun.fit_all(temp_lz[0], id_teacher, id_score, score_scale, n, same_teacher)
                    change_count[lz] = 0

            if (cost_2 > fit_fun.fit_all(a_best, id_teacher, id_score, score_scale, n, same_teacher)):
                iteration = -1
                print("ABC 全局最优发生变化")
                a_best = copy.deepcopy(ABC_lzs[lz])
        # print('\n')
        a_diversity = tools.Diversity.Diversity(ABC_lzs)


# ======================================================================================
        diversity_list.append((g_diversity+d_diversity+a_diversity)/3)
        g_best_val = fit_fun.fit_all(g_best, id_teacher, id_score, score_scale, n, same_teacher)
        d_best_val = fit_fun.fit_all(d_best, id_teacher, id_score, score_scale, n, same_teacher)
        a_best_val = fit_fun.fit_all(a_best, id_teacher, id_score, score_scale, n, same_teacher)
        temp = max(g_best_val,max(d_best_val,a_best_val))
        if g_best_val == temp:
            boost_best = copy.deepcopy(g_best)
            g_best = copy.deepcopy(boost_best)
            d_best = copy.deepcopy(boost_best)
            a_best = copy.deepcopy(boost_best)
            boost_best_list.append(g_best_val)

        elif d_best_val == temp:
            boost_best = copy.deepcopy(d_best)
            g_best = copy.deepcopy(boost_best)
            d_best = copy.deepcopy(boost_best)
            a_best = copy.deepcopy(boost_best)
            boost_best_list.append(d_best_val)
        else:
            boost_best = copy.deepcopy(a_best)
            g_best = copy.deepcopy(boost_best)
            d_best = copy.deepcopy(boost_best)
            a_best = copy.deepcopy(boost_best)
            boost_best_list.append(a_best_val)

        print("--b_best_val__:",temp)

        # 挑出没有作为答辩老师的老师
        rest_teachers = []
        for group in boost_best:
            for i in list(group[0].keys())[:-1]:
                if i not in group[0]['teachers'] and i not in no_dabian:
                    rest_teachers.append(i)
        print('rest_teachers', rest_teachers)
        # print('advance:',advance)
        # print('g_advance:',g_advance)
        # print('pre best:', pre_best)
        # print('next best:', g_best_list[-1])
        # 取没有当人答辩老师的老师来补答辩老师不足的组
        for group in boost_best:
            for no_d in no_dabian:
                if no_d in group[0]['teachers']:
                    group[0]['teachers'].pop(group[0]['teachers'].index(no_d))
                    print('no_d',no_d)
            re_time = 0
            while len(group[0]['teachers']) < teachers:
                print('rest_teachers2', rest_teachers)
                re_time += 1
                if re_time > 10:
                    break
                if len(rest_teachers) == 0:
                    continue
                random_num = random.randint(0, len(rest_teachers) - 1)
                if rest_teachers[random_num] not in list(group[0].keys())[:-1] and bool(1-if_union(teacher_statu[
                    rest_teachers[random_num]], group[1])) and bool(1-if_union(res_initial.teacher_teacher[
                    rest_teachers[random_num]], group[0]['teachers']))and rest_teachers[random_num] not in no_dabian:
                    print(res_initial.teacher_teacher[rest_teachers[random_num]])
                    print(boost_best.index(group), rest_teachers[random_num])
                    group[0]['teachers'].append(rest_teachers[random_num])
                    rest_teachers.pop(random_num)

        iteration += 1
        if iteration >= (accuracy_level * 50):  # 当迭代超过一定次数没有进步时退出迭代
            break

    # 2019/10/24
    # if fit_fun.fit_all(boost_best, id_teacher, id_score, score_scale, n, same_teacher) < 0:
    #     print("过多老师在同组答辩")
    #     return -3

    # # 打印图表
    # plot_scale(begin_best, 0,id_score)
    # plot_scale(g_best, 1,id_score)
    # plt.figure()
    # #print(len(g_best_list))
    # #print(g_best_list)
    # plt.plot(g_best_list)
    # plt.xlabel("iter_num")
    # plt.ylabel("cost (-fit)")
    # plt.show()

    web_ans = {}
    # web_ans["status"] = 200
    for index, i in enumerate(boost_best):
        web_ans[index] = {}
        web_ans[index]["dabian_teachers"] = boost_best[index][0]["teachers"]
        web_ans[index]["teachers"] = {}
        sum = 0
        for teacher, stu_num in tuple(boost_best[index][0].items())[:-1]:
            web_ans[index]["teachers"][teacher] = boost_best[index][1][sum:sum + stu_num]
            sum += stu_num

    # 打印图表
    # plot_scale(begin_best, 0,id_score)

    # plt.figure()
    # #print(len(g_best_list))
    # #print(g_best_list)

    # plt.xlabel(u"迭代次数",fontproperties=font_set)
    # plt.ylabel(u"适应度",fontproperties=font_set)
    # plt.ylim(0.8,1)
    # plt.yticks(0.9,1,0.01)
    # print("===="*20)
    # print(g_best_list)
    # print("===="*20)
    # print(g_best_list[-1])




    # 合并答辩老师
    head = 2
    end = 0
    # wb = load_workbook(r'C:\Users\44540\Desktop\结果表格.xlsx')
    # wb = load_workbook(r'files\结果表格.xlsx')
    #
    # ws = wb.active
    # for i in g_best:
    #     end = head + len(i[1]) - 1
    #     ws.merge_cells(start_row=head, start_column=6, end_row=end, end_column=6)
    #     ws.merge_cells(start_row=head, start_column=1, end_row=end, end_column=1)
    #     tem = 'A' + str(head)
    #     ws[tem].alignment = Alignment(horizontal='justify', vertical='center')
    #     ws[tem] = "第" + str(g_best.index(i)) + "组：\n" + str(len(i[1])) +"人"
    #     tem = 'A' + str(end+1)
    #     ws[tem] = " "
    #     tem = 'F'+str(head)
    #     ws[tem].alignment = Alignment(horizontal='justify', vertical='center')
    #     head = end + 2
    # # wb.save(r'C:\Users\44540\Desktop\结果表格.xlsx')
    # wb.save(r'files\结果表格.xlsx')

    web_ans["id_teacher"] = id_teacher
    web_ans["id_score"] = id_score
    web_ans = json.dumps(web_ans, ensure_ascii=False)
    # ##print(web_ans)
    # return web_ans,g_best_list[-1]
    # print(g_best)
    # for i in g_best:
    #     print(len(i[1]))

    print(boost_best_list[-1])
    for group in boost_best:
        print(group)
        print(len(group[1]))
    # print(diversity_list)
    # 将结果写到csv
    tools.resultToCSV.to_CSV(boost_best, "BOOST", id_score, id_teacher,id_name,id_theme)
    # 画出迭代过程中多样性的图
    tools.plot_fun.plot_diversity(diversity_list, "BOOST")
    # 画出迭代过程中适应度的图
    tools.plot_fun.plot_fitness(g_best_list, "BOOST")
    # //n个组的分布图
    tools.plot_fun.plot_scale(boost_best, "BOOST", id_score)

    a, b = tools.fit_fun.show_fit(boost_best, id_teacher, id_score, score_scale, n, same_teacher)
    # print("fit:")
    # print(a+b)

    return web_ans,a+b,diversity_list,boost_best_list

# def begin(n=4, x=2, n_groups=50, teachers=4, accuracy_level=2, clash_teacher=[], same_teacher=[], same_teacher_p=[]):

# 计算100次独立实验算平均、最大、最小适应度。
# hundred_time = []
# time = 0
# while(time<100):
#     try:
#         a,b,c,d = begin(n=4, accuracy_level=2, teachers=4,rd=0.5)
#         hundred_time.append(b)
#         time+=1
#         # print(len(b))
#         print("---",time,"---")
#     except Exception as e:
#         pass
#     finally:
#         pass
# print(hundred_time)
# print(max(hundred_time))
# print(min(hundred_time))
# print(sum(hundred_time)/100)
# with open(r"../output/statistic_data.txt","a+") as f:
#     f.write("BOOSTI(+ABC) 100 DIVERSITY:"+'\n' + json.dumps(hundred_time) +'\n')
#     print("成功写入！")

#
a,b,c,d = begin(n=7, accuracy_level=1,teachers=3,clash_teacher=[['陈昭炯','白清源','叶东毅'],['谢丽聪','陈昭炯']],no_dabian=['吴英杰','吴运兵','于元隆','郭昆','牛玉贞'])
print(d[-1])
print(d)
# a,b = begin(n=4, accuracy_level=4,teachers=4)
# print(b)

# 跑10次取一次话折线图
# with open(r"GA_bianyi.txt","a+") as f:
#     for i in range(10):
#         try:
#             a,b = begin(n=4, accuracy_level=2, teachers=4, rd=0.3)
#             # print(b)
#             f.write(json.dumps(b,ensure_ascii=False)+'\n')
#         finally:
#             pass

# for j in range(20):真
#     try:
#         x_index = []
#         y = []
#         for i in range(40):
#             print(i)
#             x_index.append(i / 40)
#             a, b = begin(n=4, accuracy_level=2, teachers=4, rd=i / 40)
#             y.append(b)
#
#         print(x_index)
#         print(y)
#         plt.figure()
#         plt.plot(x_index, y)
#         plt.show()
#     except Exception as e:
#         pass
#     finally:
#         with open("./rand.txt",'a+') as f:
#             f.write(json.dumps(y))
#             f.write("\n")
#
# for i in range(100):
#     print(begin(n=4, accuracy_level=2,teachers=4))
# print(begin(n=33, accuracy_level=1))




