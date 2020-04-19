
import openpyxl
import random


def random_pick(probabilities):     # 轮盘赌
    x = random.uniform(0, 1)
    cumulative_probability = 0.0
    for item_probability in probabilities:
        cumulative_probability += item_probability
        if x < cumulative_probability:
            y = random.uniform(1.5 + 0.5*probabilities.index(item_probability), 2.0 + 0.5*probabilities.index(item_probability))
            return y


def createData(total=111,a=0.1,b=0.3,c=0.5,d=0.2,e=0.1):
    '''
    在原始学生教师对应关系上创建新的绩点分布 （只改变绩点）
    :param total: 总学生数
    :param a: 1.5-2 绩点比例
    :param b: 2-2.5
    :param c: 2.5-3
    :param d: 3-3.5
    :return:
    '''
    wb = openpyxl.load_workbook(r'../input_data/new_data.xlsx')
    sh = wb["Sheet1"]
    # print(sh.cell(row=1,column=2).value)
    # sh.cell(row=1,column=4,value="111")
    teacher_student = {'A老师': [], 'B老师': [], 'C老师': [], 'D老师': [], 'E老师': [], 'F老师': [], 'G老师': [],
                       'H老师': [], 'I老师': [], 'J老师': [], 'K老师': [], 'L老师': [], 'M老师': [], 'N老师': [],
                       'O老师': [], 'P老师': [], 'Q老师': [], 'R老师': [], 'S老师': [], 'T老师': [], 'O老师': [],
                       'P老师': [], 'Q老师': [], 'R老师': [], 'S老师': [], 'T老师': [], 'U老师': [], 'V老师': [],
                       'W老师': [], 'X老师': [], 'Y老师': [], 'Z老师': [], 'Aa老师': [], 'Ab老师': [], 'Ac老师': [],
                       'Ad老师': [], 'Ae老师': [], 'Ag老师': []}

    # total = 111
    score = []
    score.append(a)
    score.append(b)
    score.append(c)
    score.append(d)
    score.append(e)
    for teacher in teacher_student.keys():
        for i in range(random.randint(2, 6)):
            teacher_student[teacher].append(random_pick(score))
    print(teacher_student)
    j = 0
    count = 0
    for teacher in teacher_student.keys():
        for i in teacher_student[teacher]:
            count += 1
            sh.cell(row=j + 2,column=1,value=count)
            sh.cell(row=j + 2, column=2, value=i)
            sh.cell(row=j + 2, column=3, value=teacher)
            j += 1

    # for i in range(111):
    # #     rand_list.append(random.uniform(1,4))
    # # print(rand_list)
    #     sh.cell(row=i+2,column=2,value=random.uniform(1,4))
    wb.save(r'../input_data/new_data.xlsx')
createData()
# import openpyxl
# import random
#
#
# def createData(total=111,a=0.1,b=0.22,c=0.4,d=0.2):
#     '''
#     在原始学生教师对应关系上创建新的绩点分布 （只改变绩点）
#     :param total: 总学生数
#     :param a: 1.5-2 绩点比例
#     :param b: 2-2.5
#     :param c: 2.5-3
#     :param d: 3-3.5
#     :return:
#     '''
#     wb = openpyxl.load_workbook(r'../input_data/new_data.xlsx')
#     sh = wb["Sheet1"]
#     # print(sh.cell(row=1,column=2).value)
#     # sh.cell(row=1,column=4,value="111")
#     total = 111
#
#     a = 0.1
#     b = 0.22
#     c = 0.4
#     d = 0.2
#
#     a_num = round(a * total)
#     b_num = round(b * total)
#     c_num = round(c * total)
#     d_num = round(d * total)
#     e_num = total - a_num - b_num - c_num - d_num
#     for i in range(a_num):
#         print(i + 2)
#         sh.cell(row=i + 2, column=2, value=random.uniform(1.5, 2))
#     for i in range(b_num):
#         print(i + 2 + a_num)
#         sh.cell(row=i + a_num + 2, column=2, value=random.uniform(2, 2.5))
#     for i in range(c_num):
#         print(i + a_num + b_num + 2)
#         sh.cell(row=i + a_num + b_num + 2, column=2, value=random.uniform(2.5, 3))
#     for i in range(d_num):
#         print(i + a_num + b_num + 2 + c_num)
#         sh.cell(row=i + a_num + b_num + 2 + c_num, column=2, value=random.uniform(3, 3.5))
#     for i in range(e_num):
#         print(i + a_num + b_num + 2 + c_num + d_num)
#         sh.cell(row=i + a_num + b_num + 2 + c_num + d_num, column=2, value=random.uniform(3.5, 4))
#     # for i in range(111):
#     # #     rand_list.append(random.uniform(1,4))
#     # # print(rand_list)
#     #     sh.cell(row=i+2,column=2,value=random.uniform(1,4))
#     wb.save(r'../input_data/new_data.xlsx')
# createData()